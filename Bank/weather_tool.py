
import pandas as pd
from meteostat import Point, Hourly
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
import matplotlib.pyplot as plt
import os
from geopy.geocoders import Nominatim
import openai
import os
from dotenv import load_dotenv
load_dotenv()

openai.api_key = os.getenv("OPENAI_API_KEY")

def generate_weather_summary(zip_code: str = "20143", reference_temp: float = 85.0, facility_type: str = "Hyperscale Data Center"):
    # === USER INPUTS ===
    use_zip = True  # Set to True to use ZIP code; False to enter lat/lon directly
    # reference_temp and facility_type are now parameters

    if use_zip:
        geolocator = Nominatim(user_agent="weather_tool")
        location = geolocator.geocode({"postalcode": zip_code, "country": "US"})
        if location is None:
            print(f"ZIP code '{zip_code}' not found via geolocation. Falling back to default ZIP '10001' (New York, NY).")
            zip_code_fallback = "10001"
            location = geolocator.geocode({"postalcode": zip_code_fallback, "country": "US"})
            if location is None:
                raise ValueError("Fallback ZIP code also failed. Please check your input.")
            zip_code = zip_code_fallback
        lat, lon = location.latitude, location.longitude
        location_name = zip_code
    else:
        lat = 48.8566      # Example: Paris
        lon = 2.3522
        location_name = "Paris_FR"

    location = Point(lat, lon)

    # === FETCH WEATHER DATA ===
    start = datetime(datetime.now().year - 20, 1, 1)
    end = datetime(datetime.now().year - 1, 12, 31)
    data = Hourly(location, start, end)
    df = data.fetch()
    df = df.reset_index()

    datetime_col_check = None
    for possible_col in ['time', 'date']:
        if possible_col in df.columns and pd.api.types.is_datetime64_any_dtype(df[possible_col]):
            datetime_col_check = possible_col
            break

    if datetime_col_check is None or df.empty or df[datetime_col_check].dt.year.nunique() < 10:
        print("Hourly data incomplete. Falling back to daily data.")
        from meteostat import Daily
        data = Daily(location, start, end)
        df = data.fetch().reset_index()
        data_granularity = "Daily"
        for possible_col in ['time', 'date']:
            if possible_col in df.columns and pd.api.types.is_datetime64_any_dtype(df[possible_col]):
                datetime_col_check = possible_col
                break
        if datetime_col_check is None:
            raise ValueError("No valid datetime column ('time' or 'date') found in fallback daily data.")
    else:
        data_granularity = "Hourly"

    # Reassign datetime_col after fallback
    datetime_col = datetime_col_check

    df['year'] = df[datetime_col].dt.year
    df['month'] = df[datetime_col].dt.month
    df['day'] = df[datetime_col].dt.day

    # Add 'season' column
    def get_season(month):
        if month in [12, 1, 2]:
            return "Winter"
        elif month in [3, 4, 5]:
            return "Spring"
        elif month in [6, 7, 8]:
            return "Summer"
        else:
            return "Fall"

    df['season'] = df['month'].apply(get_season)

    if datetime_col == 'time':
        is_hourly = True
    else:
        is_hourly = False

    if is_hourly:
        df['hour'] = df[datetime_col].dt.hour
    else:
        df['hour'] = 12  # default midday hour for daily

    df['temp_f'] = df['temp'] * 9/5 + 32
    df = df[df['temp_f'].notna()]

    # === CALCULATE HOURLY TEMP IN F ===
    df['temp_f'] = df['temp'] * 9/5 + 32

    # Drop rows with missing temperature data
    df = df[df['temp_f'].notna()]

    # === ANALYTICS ===
    summary = df.groupby('year')['temp_f'].agg(['max', 'min', 'mean', 'std']).reset_index()
    summary['five_year_avg'] = summary['mean'].rolling(window=5).mean()
    summary['delta'] = summary['max'] - summary['min']

    # Calculate hours above reference temperature per year
    df['above_ref'] = df['temp_f'] > reference_temp
    hours_above_ref = df.groupby('year')['above_ref'].sum()
    summary['above_ref'] = summary['year'].map(hours_above_ref).fillna(0).astype(int)

    # print("Reference Temp:", reference_temp)
    # print("Sample temps:\n", df['temp_f'].head())
    # print("Hours above reference per year:\n", summary[['year', 'above_ref']])
    # print("Total hours per year:\n", df.groupby('year').size())

    # Calculate total hours per year
    total_hours_per_year = df.groupby('year').size()

    # Calculate percentage hours above reference temperature
    summary['pct_above_ref'] = summary['year'].map(
        lambda y: (summary.loc[summary['year'] == y, 'above_ref'].values[0] / total_hours_per_year[y]) * 100 if y in total_hours_per_year else 0
    )

    # Identify date of max and min temperature for each year
    max_dates = df.loc[df.groupby('year')['temp_f'].idxmax()][['year', 'time']].set_index('year')
    min_dates = df.loc[df.groupby('year')['temp_f'].idxmin()][['year', 'time']].set_index('year')
    summary['max_temp_date'] = summary['year'].map(max_dates['time']).dt.strftime('%Y-%m-%d')
    summary['min_temp_date'] = summary['year'].map(min_dates['time']).dt.strftime('%Y-%m-%d')

    # === SEASONAL ANALYTICS ===
    seasonal_summary = df.groupby(['year', 'season'])['temp_f'].agg(['max', 'min', 'mean', 'std']).reset_index()

    # === CREATE EXCEL FILE ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Weather Summary"

    # Add summary row at the top
    overall_avg_mean_temp = round(summary['mean'].mean(), 1)
    year_with_highest_temp = int(summary.loc[summary['max'].idxmax(), 'year'])
    total_hours_above_ref = summary['above_ref'].sum()

    ws.append(["Overall Avg Mean Temp (F):", overall_avg_mean_temp])
    ws.append(["Year with Highest Temp:", year_with_highest_temp])
    ws.append(["Total Hours > Ref Temp:", total_hours_above_ref])
    ws.append([])  # Empty row
    # Note on data granularity and missing years
    ws.append(["Data Granularity Used:", data_granularity])

    headers = [
        "Year", "Max Temp (F)", "Min Temp (F)", "Mean Temp (F)", "Std Dev (F)",
        "Max Temp Date", "Min Temp Date", f"Hrs > {reference_temp}F",
        "% Hours > Ref Temp", "5-Year Avg Temp", "Max-Min Temp Delta"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)

    # Bold header row
    for col in ws.iter_cols(min_row=5, max_row=5, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = Font(bold=True)

    # Data rows start at row 6
    for i, row in enumerate(summary.itertuples(index=True), start=6):
        ws.cell(row=i, column=1, value=row.year)
        ws.cell(row=i, column=2, value=round(row.max, 2))
        ws.cell(row=i, column=3, value=round(row.min, 2))
        ws.cell(row=i, column=4, value=round(row.mean, 2))
        ws.cell(row=i, column=5, value=round(row.std, 2))
        ws.cell(row=i, column=6, value=row.max_temp_date)
        ws.cell(row=i, column=7, value=row.min_temp_date)
        ws.cell(row=i, column=8, value=row.above_ref)
        ws.cell(row=i, column=9, value=round(row.pct_above_ref, 2))
        ws.cell(
            row=i,
            column=10,
            value=round(summary.loc[row.Index, 'five_year_avg'], 2) if pd.notna(summary.loc[row.Index, 'five_year_avg']) else None
        )
        ws.cell(
            row=i,
            column=11,
            value=round(summary.loc[row.Index, 'delta'], 2) if pd.notna(summary.loc[row.Index, 'delta']) else None
        )

    ws.freeze_panes = ws['A6']

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Ensure all columns (including new columns) are formatted for all rows
    for row in ws.iter_rows(min_row=5, max_row=5+len(summary), min_col=1, max_col=12):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Conditional formatting to highlight max and min temps
    max_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
    min_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green

    max_temp_col = 2  # B
    min_temp_col = 3  # C
    data_start_row = 6
    data_end_row = 5 + len(summary)

    for row in range(data_start_row, data_end_row + 1):
        max_cell = ws.cell(row=row, column=max_temp_col)
        min_cell = ws.cell(row=row, column=min_temp_col)
        max_cell.fill = max_fill
        min_cell.fill = min_fill

    chart_data_start = 6
    chart_data_end = 5 + len(summary)

    # === LINE CHART ===
    chart = LineChart()
    chart.title = "Temperature Trends"
    chart.y_axis.title = 'Degrees (F)'
    chart.x_axis.title = 'Year'

    chart.add_data(Reference(ws, min_col=2, max_col=4, min_row=chart_data_start - 1, max_row=chart_data_end), titles_from_data=True)
    from openpyxl.chart.series import SeriesLabel
    chart.series[0].title = SeriesLabel(v="Max Temp")
    chart.series[1].title = SeriesLabel(v="Min Temp")
    chart.series[2].title = SeriesLabel(v="Mean Temp")
    chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end))
    ws.add_chart(chart, f"A{chart_data_end + 3}")

    # === BAR CHART for % Hours > Ref Temp ===
    bar_chart = BarChart()
    bar_chart.title = f"% Hours > {reference_temp}F Per Year"
    bar_chart.y_axis.title = '% Hours > Ref Temp'
    bar_chart.x_axis.title = 'Year'

    bar_chart.add_data(Reference(ws, min_col=9, max_col=9, min_row=chart_data_start - 1, max_row=chart_data_end), titles_from_data=True)
    bar_chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end))
    bar_chart.shape = 4
    ws.add_chart(bar_chart, f"A{chart_data_end + 20}")

    # Create seasonal charts
    season_colors = {"Winter": "blue", "Spring": "green", "Summer": "red", "Fall": "orange"}
    season_order = ["Winter", "Spring", "Summer", "Fall"]

    # Add new worksheet for seasonal charts
    season_ws = wb.create_sheet(title="Seasonal Trends")
    season_ws.append(["Seasonal Temperature Trends (2005–2024)"])
    row_cursor = 2

    for season in season_order:
        data = seasonal_summary[seasonal_summary['season'] == season]

        # Write season label
        season_ws.append([f"{season}"])
        season_ws.append(["Year", "Mean Temp (F)"])
        season_ws[f"A{row_cursor}"].font = Font(bold=True)
        season_ws[f"A{row_cursor+1}"].font = Font(bold=True)
        row_cursor += 2

        for i, row in data.iterrows():
            season_ws.append([int(row['year']), round(row['mean'], 2)])

        # Create chart
        chart = LineChart()
        chart.title = f"{season} Mean Temps"
        chart.y_axis.title = 'Mean Temperature (°F)'
        chart.x_axis.title = 'Year'

        # Enhance label styling
        chart.title.txPr = None  # Reset title style
        chart.style = 13  # Apply a built-in style for better visuals
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorTickMark = "in"
        chart.y_axis.majorTickMark = "in"
        chart.legend = None  # Remove legend for cleaner look

        start_row = row_cursor
        end_row = start_row + len(data) - 1

        data_ref = Reference(season_ws, min_col=2, min_row=start_row, max_row=end_row)
        categories = Reference(season_ws, min_col=1, min_row=start_row, max_row=end_row)

        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 15
        season_ws.add_chart(chart, f"D{row_cursor}")

        row_cursor = end_row + 3

    # === MONTHLY AVG TEMP BY YEAR ===
    monthly_avg = df.groupby(['year', 'month'])['temp_f'].mean().reset_index()
    monthly_ws = wb.create_sheet(title="Monthly Trends")
    monthly_ws.append(["Year", "Month", "Mean Temp (°F)"])

    for row in monthly_avg.itertuples(index=False):
        monthly_ws.append([int(row.year), int(row.month), round(row.temp_f, 2)])

    # Create a line chart showing average monthly temp for the most recent year
    monthly_chart = LineChart()
    monthly_chart.title = "Monthly Avg Temp (Most Recent Year)"
    monthly_chart.y_axis.title = 'Temp (°F)'
    monthly_chart.x_axis.title = 'Month'
    monthly_chart.style = 13

    latest_year = monthly_avg['year'].max()
    recent_data = monthly_avg[monthly_avg['year'] == latest_year]
    start_row = 2 + len(monthly_avg)  # Ensure chart does not overlap data

    for i, row in enumerate(recent_data.itertuples(index=False), start=start_row):
        monthly_ws.append([row.year, row.month, round(row.temp_f, 2)])

    data_ref = Reference(monthly_ws, min_col=3, min_row=start_row, max_row=start_row + 11)
    categories = Reference(monthly_ws, min_col=2, min_row=start_row, max_row=start_row + 11)
    monthly_chart.add_data(data_ref, titles_from_data=False)
    from openpyxl.chart.series import SeriesLabel
    monthly_chart.series[0].title = SeriesLabel(v=f"{latest_year}")
    monthly_chart.set_categories(categories)
    monthly_ws.add_chart(monthly_chart, "E2")


    # Diagnostic print just before export
    # print("Final years in summary table:", summary['year'].tolist())

    # === CHARTS (reuse chart_data_end) ===
    trend_chart = LineChart()
    trend_chart.title = "Mean Temp vs 5-Year Avg"
    trend_chart.y_axis.title = "Temperature (°F)"
    trend_chart.x_axis.title = "Year"
    trend_chart.style = 13
    trend_chart.add_data(Reference(ws, min_col=4, max_col=4, min_row=chart_data_start - 1, max_row=chart_data_end), titles_from_data=True)
    trend_chart.add_data(Reference(ws, min_col=10, max_col=10, min_row=chart_data_start, max_row=chart_data_end), titles_from_data=False)
    from openpyxl.chart.series import SeriesLabel
    trend_chart.series[1].title = SeriesLabel(v="5-Year Avg")
    trend_chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end))
    ws.add_chart(trend_chart, f"A{chart_data_end + 37}")

    delta_chart = LineChart()
    delta_chart.title = "Max vs Min Temp Delta"
    delta_chart.y_axis.title = "Δ Temp (°F)"
    delta_chart.x_axis.title = "Year"
    delta_chart.style = 13
    delta_chart.add_data(Reference(ws, min_col=11, max_col=11, min_row=chart_data_start, max_row=chart_data_end), titles_from_data=False)
    from openpyxl.chart.series import SeriesLabel
    delta_chart.series[0].title = SeriesLabel(v="Max-Min Δ")
    delta_chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end))
    ws.add_chart(delta_chart, f"A{chart_data_end + 54}")

    # Get top 10 hottest days
    top_10_hottest = df.sort_values(by='temp_f', ascending=False).head(10)
    top_10_list = top_10_hottest[[datetime_col, 'temp_f']].to_string(index=False)

    # Get top 10 coldest days
    top_10_coldest = df.sort_values(by='temp_f', ascending=True).head(10)
    top_10_coldest_list = top_10_coldest[[datetime_col, 'temp_f']].to_string(index=False)

    # Add new worksheet for Extreme Days
    extreme_ws = wb.create_sheet(title="Extreme Days")
    # Top 10 Hottest Days Section
    extreme_ws.append(["Top 10 Hottest Days"])
    extreme_ws.append(["Date", "Temperature (F)"])
    for cell in extreme_ws[1]:
        cell.font = Font(bold=True)
    for cell in extreme_ws[2]:
        cell.font = Font(bold=True)
    for idx, row in top_10_hottest.iterrows():
        date_str = row[datetime_col].strftime('%Y-%m-%d %H:%M') if pd.notna(row[datetime_col]) else ""
        extreme_ws.append([date_str, round(row['temp_f'], 2)])
    # Add an empty row before next section
    extreme_ws.append([])
    # Top 10 Coldest Days Section
    extreme_ws.append(["Top 10 Coldest Days"])
    extreme_ws.append(["Date", "Temperature (F)"])
    for cell in extreme_ws[extreme_ws.max_row - 1]:
        cell.font = Font(bold=True)
    for cell in extreme_ws[extreme_ws.max_row]:
        cell.font = Font(bold=True)
    for idx, row in top_10_coldest.iterrows():
        date_str = row[datetime_col].strftime('%Y-%m-%d %H:%M') if pd.notna(row[datetime_col]) else ""
        extreme_ws.append([date_str, round(row['temp_f'], 2)])

    # === SAVE FILE ===
    output_file = f"weather_summary_{location_name}.xlsx"
    wb.save(output_file)
    summary.to_csv(f"summary_{location_name}.csv", index=False)
    seasonal_summary.to_csv(f"seasonal_summary_{location_name}.csv", index=False)
    top_10_hottest.to_csv(f"hottest_{location_name}.csv", index=False)
    top_10_coldest.to_csv(f"coldest_{location_name}.csv", index=False)
    # print(f"Saved: {output_file} with updated sheets and charts.")
    return {
        "excel_file": output_file,
        "csv_summary": f"summary_{location_name}.csv",
        "csv_seasonal": f"seasonal_summary_{location_name}.csv",
        "csv_hottest": f"hottest_{location_name}.csv",
        "csv_coldest": f"coldest_{location_name}.csv"
    }
